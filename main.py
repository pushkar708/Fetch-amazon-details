import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import pyautogui as py
from selenium.webdriver.chrome.options import Options
import os
import tkinter as tk
from tkinter import simpledialog


# NOT WORKING RIGHT NOW, NEED TO UPDATE SELENIUM AND CHROME DRIVER ALONG WITH XPATHS

cwd = os.path.dirname(os.path.abspath(__file__))

def get_input():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Prompt for the word
    word = simpledialog.askstring("Enter the word to search", "Enter the word to search:")
    
    # Prompt for the number of pages
    count = simpledialog.askstring("Enter the number of pages you want to search", "Enter the number of pages you want to search:")
    
    # Ensure to show the main window again before quitting
    root.deiconify()
    root.destroy()  # Close the main window after input is received
    return word, int(count)


word , count = get_input()

if word and count:
    options = Options()

    # options.add_argument("--headless")

    driver = webdriver.Chrome(options=options)

    list_titles = []
    list_prices = []
    url_list = []

    try:
        for page in range(1, count+1):  # Just for one page, you can adjust as needed
            driver.get(f"https://www.amazon.in/s?k={word}&page={page}")
            time.sleep(3)

            if page==1:
                price_symbol=driver.find_elements(By.XPATH, "//span[@class='a-price-symbol']")[0].text
            else:
                price_symbol="INR"
            titles = driver.find_elements(By.XPATH, "//span[contains(@class,'a-color-base a-text-normal')]")
            product_url = driver.find_elements(By.XPATH, "//span[contains(@class,'a-color-base a-text-normal')]/parent::a")
            prices = driver.find_elements(By.XPATH, "//span[contains(@class,'a-price-whole')]")

            if len(titles) == 0 or len(prices) == 0:
                break

            for title, price, url in zip(titles, prices, product_url):
                list_titles.append(title.text)
                list_prices.append(price.text)
                url_list.append(url.get_attribute("href"))

            print(f"Page {page} Done")
    except Exception as e:
        print(f"Error occurred during scraping: {e}")
    finally:
        driver.quit()

    try:
        # Create a new Excel workbook
        wb = openpyxl.Workbook()

        # Select the active worksheet
        ws = wb.active

        # Write headings
        ws['A1'] = 'Title'
        ws['B1'] = f'Price in {price_symbol}'
        ws['C1'] = f'URL of the product'

        # Write data
        for idx, (title, price, url) in enumerate(zip(list_titles, list_prices, url_list), start=2):
            ws[f'A{idx}'] = title
            ws[f'B{idx}'] = price
            ws[f'C{idx}'] = url

        # Save the workbook
        wb.save(os.path.join(cwd,f'Details for {word}.xlsx'))

        print("Excel file generated successfully.")
        print(f"Details for {word} done, total results are {len(list_titles)}")

    except Exception as e:
        print(f"Error occurred while writing to Excel: {e}")
